"""
FRTB SA Capital Engine — Stage 2, Phase 3
inter_bucket.py: inter-bucket aggregation per MAR21.4(4).

For each (kind, risk_class):
    Delta_rc = sqrt( sum_b Kb^2 + sum_{b != c} gamma_bc * Sb * Sc )

with the MAR21.4(5) alternative-S floor applied when the discriminant is
negative: Sb* = max(min(Sb, Kb), -Kb), and Delta_rc recomputed.

Per-risk-class gamma rules, sourced from MAR21_Config_RW_Corr.xlsx:

    GIRR               MAR21.50      flat 0.5 between distinct currency buckets
    CSR non-sec        MAR21.57-60   gamma_sector matrix * gamma_rating (0.5 IG/HY);
                                     index buckets 17/18 use matrix entries directly
    CSR sec non-CTP    MAR21.70      0 between distinct buckets 1..24
    Equity             MAR21.80      rule-based: 0.15 within 1-10, 0 for 11,
                                     0.75 between 12/13, 0.45 for 12/13 vs 1-10
    Commodity          MAR21.85      0.20 within 1-10, 0 for 11
    FX                 MAR21.89      0.6 between distinct currency pairs
    Vega (all)         MAR21.95      same gamma as corresponding delta class

Correlation scenarios (MAR21.6) applied via the shared scenario_fn from Phase 2.
"""
import os
import math
import numpy as np
import pandas as pd

from capital_loader import load_sensitivity_grid, load_config
from weighted_sensitivities import compute_weighted_sensitivities
from intra_bucket import compute_intra_bucket, scenario_fn


# ── GIRR γ [MAR21.50] ─────────────────────────────────────────────────────────

def _girr_gamma(b1, b2, cfg) -> float:
    if b1 == b2:
        return 1.0
    val = cfg["GIRR_PARAMS"].set_index("parameter").loc["inter_bucket_gamma", "value"]
    return float(val)


# ── CSR NON-SEC γ [MAR21.57-60] ───────────────────────────────────────────────

_CSR_NONSEC_LABEL = {
    1: "1/9",  9: "1/9",
    2: "2/10", 10: "2/10",
    3: "3/11", 11: "3/11",
    4: "4/12", 12: "4/12",
    5: "5/13", 13: "5/13",
    6: "6/14", 14: "6/14",
    7: "7/15", 15: "7/15",
    8: "8",
    16: "16",
    17: "17",
    18: "18",
}


def _csr_nonsec_gamma_matrix(cfg: dict) -> pd.DataFrame:
    df = cfg["CSR_NONSEC_GAMMA"]
    df = df[df.iloc[:, 0].isin(_CSR_NONSEC_LABEL.values())]
    mat = df.set_index(df.columns[0])
    # Some columns in the config come in as strings, some as ints — force strings
    mat.columns = [str(c) for c in mat.columns]
    mat.index = [str(i) for i in mat.index]
    return mat.apply(pd.to_numeric, errors="coerce")


def _csr_nonsec_gamma(b1: int, b2: int, cfg: dict, mat: pd.DataFrame) -> float:
    if b1 == b2:
        return 1.0
    if 16 in (b1, b2):
        return 0.0   # MAR21.55: bucket 16 not aggregated with others
    lab1 = _CSR_NONSEC_LABEL[b1]
    lab2 = _CSR_NONSEC_LABEL[b2]
    g_sector = float(mat.loc[lab1, lab2])

    # gamma_rating only applies between single-name buckets (1-15 incl. 8).
    # Indices (17/18) use matrix values directly per MAR21.60.
    single_name = lambda b: b in range(1, 16)
    if single_name(b1) and single_name(b2):
        rating_1 = "IG" if b1 <= 8 else "HY"
        rating_2 = "IG" if b2 <= 8 else "HY"
        g_rating = 1.0 if rating_1 == rating_2 else 0.5
        return g_sector * g_rating
    return g_sector


# ── CSR SEC NON-CTP γ [MAR21.70] ──────────────────────────────────────────────

def _csr_sec_nonctp_gamma(b1: int, b2: int, cfg: dict) -> float:
    if b1 == b2:
        return 1.0
    val = cfg["CSR_SEC_NONCTP_CORR"].set_index("parameter").loc[
        "inter_bucket_1_to_24", "value"]
    return float(val)


# ── EQUITY γ [MAR21.80] ───────────────────────────────────────────────────────

def _equity_gamma(b1: int, b2: int, cfg: dict) -> float:
    if b1 == b2:
        return 1.0
    p = cfg["EQUITY_CORR"].set_index("parameter")["value"]
    if 11 in (b1, b2):
        return float(p.loc["inter_bucket_11_vs_any"])
    if {b1, b2} == {12, 13}:
        return float(p.loc["inter_bucket_12_vs_13"])
    indices = {12, 13}
    single_names = set(range(1, 11))
    if (b1 in indices and b2 in single_names) or (b1 in single_names and b2 in indices):
        return float(p.loc["inter_bucket_12or13_vs_1to10"])
    if b1 in single_names and b2 in single_names:
        return float(p.loc["inter_bucket_1_to_10"])
    raise ValueError(f"Equity gamma unhandled: ({b1}, {b2})")


# ── COMMODITY γ [MAR21.85] ────────────────────────────────────────────────────

def _commodity_gamma(b1: int, b2: int, cfg: dict) -> float:
    if b1 == b2:
        return 1.0
    p = cfg["COMMODITY_CORR"].set_index("parameter")["value"]
    if 11 in (b1, b2):
        return float(p.loc["inter_bucket_11_vs_any"])
    return float(p.loc["inter_bucket_1_to_10"])


# ── FX γ [MAR21.89] ───────────────────────────────────────────────────────────

def _fx_gamma(b1, b2, cfg) -> float:
    if b1 == b2:
        return 1.0
    val = cfg["FX_PARAMS"].set_index("parameter").loc["inter_bucket_gamma", "value"]
    return float(val)


# ── CORE INTER-BUCKET AGGREGATOR ──────────────────────────────────────────────

def _delta_rc(Kbs: np.ndarray, Sbs: np.ndarray, gamma_fn, scen) -> tuple[float, bool]:
    """Compute Delta_rc and floor flag (without exposing the gamma matrix)."""
    Delta, floored, _ = _delta_rc_with_trace(Kbs, Sbs, gamma_fn, scen)
    return Delta, floored


def _delta_rc_with_trace(Kbs: np.ndarray, Sbs: np.ndarray, gamma_fn, scen,
                          linear_add_mask: np.ndarray = None
                          ) -> tuple[float, bool, np.ndarray]:
    """
    Compute Delta_rc and ALSO return the (n, n) scenario-scaled gamma matrix
    actually used (for audit / traceability output).

    linear_add_mask: optional boolean array of length n. Buckets where True are
    EXCLUDED from the sqrt aggregation and ADDED LINEARLY to the result. This
    implements MAR21.71 (CSR sec bucket 25) and similar 'no diversification'
    rules. If None, all buckets enter the sqrt formula (the standard case).
    """
    n = len(Kbs)
    if n == 0:
        return 0.0, False, np.zeros((0, 0))
    gamma_mat = np.eye(n)
    if linear_add_mask is None:
        linear_add_mask = np.zeros(n, dtype=bool)
    in_sqrt = ~linear_add_mask
    Kbs_in = Kbs[in_sqrt]
    Sbs_in = Sbs[in_sqrt]
    Kbs_lin = Kbs[linear_add_mask].sum()    # linearly-added bucket Kbs

    discr = float((Kbs_in ** 2).sum())
    # Index map from in_sqrt position to original index, to keep gamma_mat full
    in_idx = np.where(in_sqrt)[0]
    for a, i_orig in enumerate(in_idx):
        for b, j_orig in enumerate(in_idx):
            if b <= a:
                continue
            g = scen(gamma_fn(i_orig, j_orig))
            gamma_mat[i_orig, j_orig] = gamma_mat[j_orig, i_orig] = g
            discr += 2.0 * g * Sbs_in[a] * Sbs_in[b]

    if discr >= 0:
        return math.sqrt(discr) + Kbs_lin, False, gamma_mat

    Sbs_star_in = np.clip(Sbs_in, -Kbs_in, Kbs_in)
    discr2 = float((Kbs_in ** 2).sum())
    for a, i_orig in enumerate(in_idx):
        for b, j_orig in enumerate(in_idx):
            if b <= a:
                continue
            discr2 += 2.0 * gamma_mat[i_orig, j_orig] * Sbs_star_in[a] * Sbs_star_in[b]
    return math.sqrt(max(0.0, discr2)) + Kbs_lin, True, gamma_mat


# ── DISPATCH ──────────────────────────────────────────────────────────────────

def _build_gamma_fn(kind: str, risk_class: str, buckets: list, cfg: dict):
    """Return a closure gamma_fn(i, j) that looks up γ between the i-th and j-th
    bucket in `buckets`. For vega, gamma mirrors the delta class (MAR21.95)."""
    # For GIRR/FX, buckets are currencies/pairs — γ is a constant, but we still
    # honour same/different via the helper signatures.
    if risk_class == "GIRR":
        return lambda i, j: _girr_gamma(buckets[i], buckets[j], cfg)
    if risk_class == "FX":
        return lambda i, j: _fx_gamma(buckets[i], buckets[j], cfg)
    if risk_class == "CSR_NONSEC":
        mat = _csr_nonsec_gamma_matrix(cfg)
        return lambda i, j: _csr_nonsec_gamma(buckets[i], buckets[j], cfg, mat)
    if risk_class == "CSR_SEC_NONCTP":
        return lambda i, j: _csr_sec_nonctp_gamma(buckets[i], buckets[j], cfg)
    if risk_class == "EQUITY":
        return lambda i, j: _equity_gamma(buckets[i], buckets[j], cfg)
    if risk_class == "COMMODITY":
        return lambda i, j: _commodity_gamma(buckets[i], buckets[j], cfg)
    raise ValueError(f"No gamma rule for ({kind}, {risk_class})")


def _linear_add_mask(kind: str, risk_class: str, buckets: list) -> np.ndarray:
    """Return a bool mask flagging buckets that should be linearly added to the
    risk-class total instead of entering the sqrt aggregation.

    Currently only CSR_SEC_NONCTP bucket 25 per MAR21.71 ('the capital
    requirements for bucket 25 and the aggregated capital requirements for
    buckets 1 to 24 will be simply summed up'). Other 'Other' buckets
    (CSR non-sec 16, equity 11, commodity 11) use gamma=0 with all other
    buckets which puts their Kb^2 INSIDE the sqrt -- mathematically different
    from linear add. Only CSR_SEC_NONCTP 25 is explicitly outside-the-sqrt.
    """
    if risk_class == "CSR_SEC_NONCTP":
        return np.array([b == 25 for b in buckets], dtype=bool)
    return np.zeros(len(buckets), dtype=bool)


def compute_inter_bucket(phase2_df: pd.DataFrame, cfg: dict,
                         scenario: str = "medium") -> pd.DataFrame:
    """
    Aggregate Phase 2 (Kb, Sb) per (kind, risk_class) into a single Delta_rc
    (or Vega_rc) number, returned as a DataFrame one row per (kind, risk_class).
    """
    scen = scenario_fn(scenario)
    if phase2_df.empty:
        return pd.DataFrame(columns=["kind", "risk_class", "n_buckets",
                                     "sum_Kb2_sqrt", "charge",
                                     "alt_s_floor_used", "scenario"])
    rows = []
    for (kind, rc), sub in phase2_df.groupby(["kind", "risk_class"]):
        sub = sub.sort_values("bucket_id").reset_index(drop=True)
        buckets = sub["bucket_id"].tolist()
        Kbs = sub["Kb"].values.astype(float)
        Sbs = sub["Sb"].values.astype(float)
        gamma_fn = _build_gamma_fn(kind, rc, buckets, cfg)
        mask = _linear_add_mask(kind, rc, buckets)
        Delta, floored, _ = _delta_rc_with_trace(Kbs, Sbs, gamma_fn, scen, mask)
        rows.append(dict(kind=kind, risk_class=rc,
                         n_buckets=len(buckets),
                         sum_Kb2_sqrt=float(math.sqrt((Kbs ** 2).sum())),
                         charge=Delta,
                         alt_s_floor_used=floored,
                         scenario=scenario))
    return pd.DataFrame(rows).sort_values(["kind", "risk_class"]).reset_index(drop=True)


def compute_inter_bucket_with_trace(phase2_df: pd.DataFrame, cfg: dict,
                                    scenario: str = "medium"):
    """Like compute_inter_bucket but additionally returns:
        bucket_inputs_df  one row per (kind, risk_class, bucket_id) showing the
                          Sb / Kb pair that fed the inter-bucket aggregation
        gamma_blocks      list of (rc_label, bucket_labels, gamma_matrix)
                          one entry per risk class with multiple buckets
    """
    scen = scenario_fn(scenario)
    if phase2_df.empty:
        empty_summary = pd.DataFrame(columns=["kind", "risk_class", "n_buckets",
                                              "sum_Kb2_sqrt", "charge",
                                              "alt_s_floor_used", "scenario"])
        return empty_summary, pd.DataFrame(), []

    summary_rows = []
    bucket_input_rows = []
    gamma_blocks = []

    for (kind, rc), sub in phase2_df.groupby(["kind", "risk_class"]):
        sub = sub.sort_values("bucket_id").reset_index(drop=True)
        buckets = sub["bucket_id"].tolist()
        Kbs = sub["Kb"].values.astype(float)
        Sbs = sub["Sb"].values.astype(float)
        gamma_fn = _build_gamma_fn(kind, rc, buckets, cfg)
        mask = _linear_add_mask(kind, rc, buckets)
        Delta, floored, gamma_mat = _delta_rc_with_trace(Kbs, Sbs, gamma_fn, scen, mask)

        summary_rows.append(dict(
            kind=kind, risk_class=rc, n_buckets=len(buckets),
            sum_Kb2_sqrt=float(math.sqrt((Kbs ** 2).sum())),
            charge=Delta, alt_s_floor_used=floored, scenario=scenario))

        for i, b in enumerate(buckets):
            bucket_input_rows.append(dict(
                scenario=scenario, kind=kind, risk_class=rc, bucket_id=b,
                Sb=float(Sbs[i]), Kb=float(Kbs[i]),
                n_factors=int(sub["n_factors"].iloc[i])))

        if len(buckets) >= 2:
            labels = [f"{rc} | {b}" for b in buckets]
            gamma_blocks.append((f"{kind} | {rc}", labels, gamma_mat))

    summary_df = pd.DataFrame(summary_rows).sort_values(
        ["kind", "risk_class"]).reset_index(drop=True)
    inputs_df = pd.DataFrame(bucket_input_rows)
    return summary_df, inputs_df, gamma_blocks


def _write_blocks_to_sheet(blocks: list, sheet, label_prefix: str = "Risk class") -> None:
    """Stack matrix blocks vertically into a worksheet."""
    row = 1
    for block_label, labels, mat in blocks:
        sheet.cell(row=row, column=1,
                   value=f"{label_prefix}: {block_label}   (n_buckets={len(labels)})")
        row += 1
        for j, lab in enumerate(labels):
            sheet.cell(row=row, column=j + 2, value=lab)
        row += 1
        for i, lab in enumerate(labels):
            sheet.cell(row=row, column=1, value=lab)
            for j in range(len(labels)):
                sheet.cell(row=row, column=j + 2, value=float(mat[i, j]))
            row += 1
        row += 2


if __name__ == "__main__":
    from openpyxl import Workbook

    print("Loading Phase 0 + 1 + 2 for all three scenarios...")
    tidy = load_sensitivity_grid()
    cfg = load_config()
    ws = compute_weighted_sensitivities(tidy, cfg)

    for scen in ("medium", "high", "low"):
        print(f"\n=== PHASE 3 INTER-BUCKET | scenario = {scen} ===")
        p2 = compute_intra_bucket(ws, cfg, scenario=scen)
        p3 = compute_inter_bucket(p2, cfg, scenario=scen)
        print(p3.to_string(index=False,
                           formatters={"sum_Kb2_sqrt": "{:,.2f}".format,
                                       "charge": "{:,.2f}".format}))
        print(f"Sum of (delta + vega) charges across all risk classes: "
              f"{p3['charge'].sum():,.2f}")

    out_path = os.path.join(os.path.dirname(__file__), "output",
                            "phase3_inter_bucket.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    readme = pd.DataFrame([
        ("Sheet", "Purpose"),
        ("README", "This sheet."),
        ("{scen}_summary",
         "Per-risk-class inter-bucket charge: sum_Kb2_sqrt (= sqrt of sum of Kb^2 "
         "ignoring cross-terms) vs charge (= full MAR21.4(4) result with gamma)."),
        ("{scen}_bucket_inputs",
         "Per (risk_class, bucket): the Kb and Sb that fed inter-bucket aggregation. "
         "These come from Phase 2; reproduced here so the trace is self-contained."),
        ("{scen}_gamma_matrices",
         "Stacked inter-bucket gamma matrices, one block per risk class with >= 2 "
         "buckets. Already scenario-scaled per MAR21.6. Single-bucket risk classes "
         "(e.g. FX with one currency pair) are omitted because charge = Kb directly."),
        ("MAR21.4(4) formula",
         "Delta_rc = sqrt( sum_b Kb^2 + sum_{b != c} gamma_bc * Sb * Sc ). If the "
         "discriminant is negative, MAR21.4(5) alternative-S floor is applied "
         "(Sb -> clip(Sb, -Kb, Kb)) and the charge recomputed; flag in summary."),
    ], columns=["A", "B"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        readme.to_excel(xw, sheet_name="README", index=False, header=False)
        for scen in ("medium", "high", "low"):
            p2 = compute_intra_bucket(ws, cfg, scenario=scen)
            summary, inputs, gamma_blocks = compute_inter_bucket_with_trace(
                p2, cfg, scenario=scen)
            summary.to_excel(xw, sheet_name=f"{scen}_summary", index=False)
            inputs.to_excel(xw, sheet_name=f"{scen}_bucket_inputs", index=False)
            sheet = xw.book.create_sheet(f"{scen}_gamma_matrices")
            _write_blocks_to_sheet(gamma_blocks, sheet, label_prefix="Risk class")
    print(f"\nWrote: {out_path}")
