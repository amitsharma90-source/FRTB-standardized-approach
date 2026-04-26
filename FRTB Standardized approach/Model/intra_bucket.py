"""
FRTB SA Capital Engine — Stage 2, Phase 2
intra_bucket.py: intra-bucket aggregation.

For every (risk_class, bucket), compute:
    Sb = sum_k WS_k                       (signed net)
    Kb = sqrt( max(0, sum_k WS_k^2 + sum_{k != l} rho_kl * WS_k * WS_l) )

Correlation rho_kl is risk-class specific, driven entirely by the config:
    GIRR delta           MAR21.45-49   tenor matrix x curve-indicator, plus
                                       inflation vs yield = 0.40, xccy vs any = 0.
    CSR non-sec delta    MAR21.54-55   rho_name * rho_tenor * rho_basis
    CSR sec non-CTP      MAR21.68      rho_tranche * rho_tenor * rho_basis
    Equity delta         MAR21.78      same-issuer=1, cross-issuer per bucket-type
    Commodity delta      MAR21.83      rho_commodity * rho_tenor * rho_basis
    FX delta             MAR21.88      single factor per pair -> Kb = |WS|
    Vega (all classes)   MAR21.94      min(rho_optmat * rho_underlying_delta, 1)

Correlation scenarios (MAR21.6) are applied via a scenario_multiplier function
before building the matrix, so Phase 5 can re-run Phases 2-4 under low/med/high.

Spec-driven: every number below is pulled from MAR21_Config_RW_Corr.xlsx.
"""
import os
import math
import numpy as np
import pandas as pd

from capital_loader import load_config
from weighted_sensitivities import compute_weighted_sensitivities
from capital_loader import load_sensitivity_grid


# ── SCENARIO APPLICATION ──────────────────────────────────────────────────────

def scenario_fn(scenario: str):
    """Return a function that maps prescribed rho -> scenario rho [MAR21.6]."""
    if scenario == "medium":
        return lambda rho: rho
    if scenario == "high":
        return lambda rho: min(rho * 1.25, 1.0)
    if scenario == "low":
        return lambda rho: max(2.0 * rho - 1.0, 0.75 * rho)
    raise ValueError(f"Unknown scenario: {scenario}")


# ── GIRR DELTA [MAR21.45-49] ──────────────────────────────────────────────────

def _girr_tenor_corr_matrix(cfg: dict) -> pd.DataFrame:
    """Return the GIRR tenor x tenor correlation matrix with float tenor keys."""
    df = cfg["GIRR_CORR"].dropna(how="all")
    def _t(x):
        try:
            return float(str(x).strip().rstrip("Y"))
        except (TypeError, ValueError):
            return None
    # Keep only rows whose first cell is a parseable tenor
    first = df.iloc[:, 0]
    mask = first.apply(lambda v: _t(v) is not None)
    df = df.loc[mask]
    parsed_rows = [_t(t) for t in df.iloc[:, 0].tolist()]
    parsed_cols = [_t(c) for c in df.columns[1:]]
    # Same filter on columns in case of stray extras
    col_keep = [i for i, v in enumerate(parsed_cols) if v is not None]
    parsed_cols = [parsed_cols[i] for i in col_keep]
    mat = df.iloc[:, 1:].iloc[:, col_keep].apply(pd.to_numeric, errors="coerce").values
    return pd.DataFrame(mat, index=parsed_rows, columns=parsed_cols)


def _girr_rho(i, j, tenor_mat, params) -> float:
    """
    Correlation between two GIRR risk factors within the same currency bucket.
    rows i, j are Series from the ws_df with fields: tenor, under_tenor, name
    (we assume same currency = same curve in this engine; specified_ccy handling
    is done in the RW only, not here).
    Special tenor values 'inflation' and 'xccy_basis' trigger MAR21.48/49.
    """
    t_i, t_j = i["tenor"], j["tenor"]
    infl_rho = params["inflation_vs_yield"]
    xccy_rho = params["xccy_basis_vs_any"]
    # xccy basis vs anything: 0.0
    if t_i == "xccy_basis" or t_j == "xccy_basis":
        if t_i == t_j:
            return 1.0
        return xccy_rho
    # inflation vs anything (except itself): 0.4
    if t_i == "inflation" or t_j == "inflation":
        if t_i == t_j:
            return 1.0
        return infl_rho
    # both are yield curve tenors
    t_i_f, t_j_f = float(t_i), float(t_j)
    if t_i_f == t_j_f:
        return 1.0
    return float(tenor_mat.loc[t_i_f, t_j_f])


# ── CSR NON-SEC DELTA [MAR21.54-55] ───────────────────────────────────────────

def _csr_nonsec_rho_params(cfg: dict, bucket: int) -> dict:
    """Return the (rho_name_diff, rho_tenor_diff, rho_basis_diff) for a bucket."""
    df = cfg["CSR_NONSEC_CORR"].set_index("parameter")
    # Column selection depends on bucket range
    if bucket in (17, 18):
        col = "buckets_17_18"
    else:
        col = "buckets_1_15"     # covers buckets 1-15; bucket 16 handled separately
    return dict(
        rho_name_diff=float(df.loc["rho_name_diff", col]),
        rho_tenor_diff=float(df.loc["rho_tenor_diff", col]),
        rho_basis_diff=float(df.loc["rho_basis_diff", col]),
    )


def _csr_nonsec_rho(i, j, bucket_params) -> float:
    """rho_kl = rho_name * rho_tenor * rho_basis [MAR21.54-55].
    Name = Security; basis is uniformly 'yield' in Stage 1 (rho_basis = 1 always)."""
    rho_name = 1.0 if i["security"] == j["security"] else bucket_params["rho_name_diff"]
    rho_tenor = 1.0 if i["tenor"] == j["tenor"] else bucket_params["rho_tenor_diff"]
    rho_basis = 1.0    # no basis dimension emitted by Stage 1
    return rho_name * rho_tenor * rho_basis


# ── CSR SEC NON-CTP [MAR21.68] ────────────────────────────────────────────────

def _csr_sec_nonctp_params(cfg: dict) -> dict:
    df = cfg["CSR_SEC_NONCTP_CORR"].set_index("parameter")["value"]
    return dict(
        rho_tranche_diff=float(df.loc["rho_tranche_diff"]),
        rho_tenor_diff=float(df.loc["rho_tenor_diff"]),
        rho_basis_diff=float(df.loc["rho_basis_diff"]),
    )


def _csr_sec_nonctp_rho(i, j, params) -> float:
    rho_tranche = 1.0 if i["security"] == j["security"] else params["rho_tranche_diff"]
    rho_tenor = 1.0 if i["tenor"] == j["tenor"] else params["rho_tenor_diff"]
    rho_basis = 1.0
    return rho_tranche * rho_tenor * rho_basis


# ── EQUITY DELTA [MAR21.78] ───────────────────────────────────────────────────

def _equity_cross_issuer_rho(bucket: int, cfg: dict) -> float:
    """Cross-issuer spot rho for a given bucket (MAR21.78(2))."""
    eq_meta = cfg["EQUITY_RW"].set_index("bucket")
    eq_corr = cfg["EQUITY_CORR"].set_index("parameter")["value"]
    if bucket in (12, 13):
        return float(eq_corr.loc["intra_index_spot"])
    mcap = str(eq_meta.loc[bucket, "market_cap"]).strip().lower()
    econ = str(eq_meta.loc[bucket, "economy"]).strip().upper()
    key_map = {
        ("large", "EM"):  "intra_large_em_spot",
        ("large", "ADV"): "intra_large_adv_spot",
        ("small", "EM"):  "intra_small_em_spot",
        ("small", "ADV"): "intra_small_adv_spot",
    }
    key = key_map.get((mcap, econ))
    if key is None:
        # Bucket 11 "Other" or malformed row — MAR21.78(3): simple sum
        return None
    return float(eq_corr.loc[key])


def _equity_rho(i, j, cross_rho: float) -> float:
    if i["bucket"] != j["bucket"]:
        raise ValueError("Intra-bucket equity rho called on different buckets")
    if i["name"] == j["name"]:
        return 1.0          # Stage 1 emits spot only, so same issuer = 1.0
    return cross_rho


# ── COMMODITY DELTA [MAR21.83] ────────────────────────────────────────────────

def _commodity_rho(i, j, cfg: dict) -> float:
    """rho_kl = rho_commodity * rho_tenor * rho_basis [MAR21.83].
    Stage 1 emits one sensitivity per (bucket, commodity) so tenor and basis are
    always same -> rho_kl = rho_commodity.
    """
    bucket = int(i["bucket"])
    if i["name"] == j["name"]:
        return 1.0
    rw_df = cfg["COMMODITY_RW"].set_index("bucket")
    return float(rw_df.loc[bucket, "intra_corr_commodity"])


# ── FX DELTA [MAR21.88] ───────────────────────────────────────────────────────
# Only one risk factor per pair; intra-bucket rho is irrelevant (Kb = |WS|).


# ── VEGA (ALL CLASSES) [MAR21.94] ─────────────────────────────────────────────

def _vega_exp_rho(t_i: float, t_j: float, alpha: float) -> float:
    """Generic vega exponential decay rho: exp(-alpha * |t_i - t_j| / min(t_i, t_j))
    [MAR21.93]. Used for both option-maturity (rho_optmat) and, for GIRR vega,
    underlying-residual maturity (rho_und)."""
    if t_i == t_j:
        return 1.0
    return math.exp(-alpha * abs(t_i - t_j) / min(t_i, t_j))


# Backward-compatible alias (older name retained for callers/tests)
_vega_optmat_rho = _vega_exp_rho


def _vega_rho(i, j, cfg: dict, alpha: float, tenor_mat) -> float:
    """Vega corr = min(rho_optmat * rho_underlying, 1) [MAR21.93-94].

    For GIRR vega [MAR21.93]:
        rho_optmat = exp(-alpha * |T_opt_k - T_opt_l| / min(T_opt_k, T_opt_l))
        rho_und    = exp(-alpha * |T_und_k - T_und_l| / min(T_und_k, T_und_l))
        i.e. BOTH terms use the exponential decay with alpha = 1%.

    For non-GIRR vega [MAR21.94]:
        rho_optmat = same exponential as above
        rho_und    = corresponding delta correlation between the underlying risk
                     factors (e.g. equity intra-bucket spot rho).
    """
    rc = i["risk_class"]
    rho_optmat = _vega_exp_rho(float(i["tenor"]), float(j["tenor"]), alpha)
    if rc == "GIRR":
        # MAR21.93(2): rho_und is the SAME exponential decay, but on the
        # underlying-residual maturity (years AFTER the option matures).
        t_i_u, t_j_u = float(i["under_tenor"]), float(j["under_tenor"])
        rho_und = _vega_exp_rho(t_i_u, t_j_u, alpha)
    elif rc == "EQUITY":
        # Underlying is the equity spot; same bucket + same name = 1, else cross
        if i["name"] == j["name"]:
            rho_und = 1.0
        else:
            rho_und = _equity_cross_issuer_rho(int(i["bucket"]), cfg)
    else:
        # FX / CSR / Commodity vega not in current portfolio — default to 1 for
        # same-name, underlying-delta rho would otherwise apply
        rho_und = 1.0 if i.get("name") == j.get("name") else 0.0
    return min(rho_optmat * rho_und, 1.0)


# ── CORE AGGREGATOR ───────────────────────────────────────────────────────────

def _kb_sb(rows: pd.DataFrame, rho_fn, scen) -> tuple[float, float, int]:
    """
    Compute Kb and Sb for a bucket given rows (DataFrame slice with `ws`) and
    a pairwise rho function rho_fn(row_i, row_j) -> float.
    """
    Kb, Sb, n, _ = _kb_sb_with_trace(rows, rho_fn, scen)
    return Kb, Sb, n


def _kb_sb_with_trace(rows: pd.DataFrame, rho_fn, scen
                      ) -> tuple[float, float, int, np.ndarray]:
    """Same as _kb_sb but also returns the (n, n) scenario-scaled rho matrix
    actually used in the Kb computation, for audit / traceability output."""
    n = len(rows)
    if n == 0:
        return 0.0, 0.0, 0, np.zeros((0, 0))
    ws = rows["ws"].values.astype(float)
    Sb = float(ws.sum())
    rho_mat = np.eye(n)
    if n == 1:
        return abs(ws[0]), Sb, 1, rho_mat
    kb2 = float((ws ** 2).sum())
    records = rows.to_dict("records")
    for a in range(n):
        for b in range(a + 1, n):
            rho = scen(rho_fn(records[a], records[b]))
            rho_mat[a, b] = rho_mat[b, a] = rho
            kb2 += 2.0 * rho * ws[a] * ws[b]
    Kb = math.sqrt(max(0.0, kb2))
    return Kb, Sb, n, rho_mat


def _kb_simple_sum(rows: pd.DataFrame) -> tuple[float, float, int]:
    """Bucket 'Other' rule: Kb = sum |WS|, Sb = sum WS (no diversification)."""
    if len(rows) == 0:
        return 0.0, 0.0, 0
    ws = rows["ws"].values.astype(float)
    return float(np.abs(ws).sum()), float(ws.sum()), len(rows)


# ── DISPATCH BY RISK CLASS ────────────────────────────────────────────────────

def _bucket_rho_fn(rows: pd.DataFrame, cfg: dict):
    """Return (rho_fn, mode) for a bucket. mode='pairwise' for normal Kb formula,
    mode='simple_sum' for the special MAR21 'Other' buckets."""
    rc = rows["risk_class"].iloc[0]
    kind = rows["kind"].iloc[0]
    if kind == "delta":
        if rc == "GIRR":
            tenor_mat = _girr_tenor_corr_matrix(cfg)
            params = cfg["GIRR_PARAMS"].set_index("parameter")["value"].astype(float).to_dict()
            return lambda a, b: _girr_rho(a, b, tenor_mat, params), "pairwise"
        if rc == "CSR_NONSEC":
            bucket = int(rows["bucket"].iloc[0])
            if bucket == 16:
                return None, "simple_sum"
            bp = _csr_nonsec_rho_params(cfg, bucket)
            return lambda a, b: _csr_nonsec_rho(a, b, bp), "pairwise"
        if rc == "CSR_SEC_NONCTP":
            bucket = int(rows["bucket"].iloc[0])
            if bucket == 25:
                return None, "simple_sum"
            bp = _csr_sec_nonctp_params(cfg)
            return lambda a, b: _csr_sec_nonctp_rho(a, b, bp), "pairwise"
        if rc == "CSR_SEC_CTP":
            # MAR21.58-61: CTP buckets reuse non-sec bucket structure with a
            # MODIFIED rho_basis (99.00% instead of 99.90% per MAR21.60(1)).
            # Stage 1 currently emits no CSR_SEC_CTP rows; this branch raises
            # explicitly so the gap is loud rather than silent if CTP is added.
            raise NotImplementedError(
                "CSR_SEC_CTP intra-bucket aggregation not implemented. "
                "MAR21.58-61 requires a modified rho_basis = 99.00% and reuse "
                "of CSR non-sec bucket/correlation structure (excluding indices). "
                "Add a dedicated rho function and route here when the portfolio "
                "starts emitting CSR_SEC_CTP sensitivities.")
        if rc == "EQUITY":
            bucket = int(rows["bucket"].iloc[0])
            cross = _equity_cross_issuer_rho(bucket, cfg)
            if cross is None:
                return None, "simple_sum"
            return lambda a, b: _equity_rho(a, b, cross), "pairwise"
        if rc == "COMMODITY":
            bucket = int(rows["bucket"].iloc[0])
            if bucket == 11:
                return None, "simple_sum"
            return lambda a, b: _commodity_rho(a, b, cfg), "pairwise"
        if rc == "FX":
            return lambda a, b: 1.0, "pairwise"
    if kind == "vega":
        alpha = float(cfg["MODEL_PARAMS"].set_index("parameter").loc["vega_alpha", "value"])
        tenor_mat = _girr_tenor_corr_matrix(cfg)
        return lambda a, b: _vega_rho(a, b, cfg, alpha, tenor_mat), "pairwise"
    raise ValueError(f"Unsupported (kind, risk_class) = ({kind}, {rc})")


def _aggregate_one_bucket(rows: pd.DataFrame, cfg: dict, scen) -> tuple[float, float, int]:
    """Aggregate a single (risk_class, bucket, kind) slice into (Kb, Sb, n)."""
    rho_fn, mode = _bucket_rho_fn(rows, cfg)
    if mode == "simple_sum":
        return _kb_simple_sum(rows)
    return _kb_sb(rows, rho_fn, scen)


def _aggregate_one_bucket_with_trace(rows: pd.DataFrame, cfg: dict, scen
                                     ) -> tuple[float, float, int, np.ndarray,
                                                str, pd.DataFrame]:
    """Aggregate a bucket and return everything needed for the audit trail:
        Kb, Sb, n_distinct_factors, rho_matrix (over distinct factors),
        mode ('pairwise' or 'simple_sum'), aggregated_factors_df.

    Tidy rows are first collapsed to distinct risk factors via
    _aggregate_to_distinct_factors so the rho matrix has one row per MAR21
    risk factor — not one per (position x factor) pair. Kb is unchanged
    (collapsing same-factor rows is mathematically equivalent under rho=1).
    """
    rho_fn, mode = _bucket_rho_fn(rows, cfg)
    if mode == "simple_sum":
        Kb, Sb, n = _kb_simple_sum(rows)
        # Even for simple_sum, expose the aggregated factor view for audit
        agg = _aggregate_to_distinct_factors(rows)
        return Kb, Sb, len(agg), np.eye(len(agg)), "simple_sum", agg
    agg = _aggregate_to_distinct_factors(rows)
    Kb, Sb, n_distinct, mat = _kb_sb_with_trace(agg, rho_fn, scen)
    return Kb, Sb, n_distinct, mat, "pairwise", agg


# ── TOP-LEVEL ─────────────────────────────────────────────────────────────────

def compute_intra_bucket(ws_df: pd.DataFrame, cfg: dict,
                         scenario: str = "medium") -> pd.DataFrame:
    """
    Return one row per (kind, risk_class, bucket_id) with columns:
        kind, risk_class, bucket_id, n_factors, Sb, Kb, scenario
    bucket_id is the currency for GIRR (and FX pair for FX), else the integer bucket.
    Curvature rows are skipped — Phase 4 handles them.
    """
    scen = scenario_fn(scenario)
    mask = ws_df["kind"].isin(["delta", "vega"])
    df = ws_df.loc[mask].copy()
    if df.empty:
        return pd.DataFrame(columns=["kind", "risk_class", "bucket_id",
                                     "n_factors", "Sb", "Kb", "scenario"])
    # Bucket key: integer bucket for most; ccy for GIRR; name for FX
    def _bucket_id(row):
        if row["risk_class"] == "GIRR":
            return row["ccy"]
        if row["risk_class"] == "FX":
            return row["name"]
        return int(row["bucket"])
    df["bucket_id"] = df.apply(_bucket_id, axis=1)

    results = []
    for (kind, rc, bid), sub in df.groupby(["kind", "risk_class", "bucket_id"]):
        Kb, Sb, n = _aggregate_one_bucket(sub, cfg, scen)
        results.append(dict(kind=kind, risk_class=rc, bucket_id=bid,
                            n_factors=n, Sb=Sb, Kb=Kb, scenario=scenario))
    return pd.DataFrame(results).sort_values(
        ["kind", "risk_class", "bucket_id"]).reset_index(drop=True)


def _factor_label(row) -> str:
    """Compact, readable factor identifier used in trace sheets.
    Combines security/issuer (where relevant) with the column name so the row
    is uniquely identifiable inside its bucket."""
    rc = row["risk_class"]
    base = row["column"]
    if rc.startswith("CSR"):
        return f"{base}  [{row['security']}]"
    return base


def _factor_identity(row) -> tuple:
    """Identity tuple that defines a distinct MAR21 risk factor.

    Two tidy rows refer to the same risk factor (so rho_kl = 1) iff their
    identity tuples are equal. Used to collapse N positions into the K distinct
    risk factors they expose, before building the rho matrix.

    GIRR / EQ / COMM / FX / Vega: column alone identifies the risk factor.
    CSR non-sec / sec: (column, security) — issuer/tranche isn't in column.
    """
    rc = row["risk_class"]
    if rc.startswith("CSR"):
        return (row["column"], row["security"])
    return (row["column"],)


def _aggregate_to_distinct_factors(rows: pd.DataFrame) -> pd.DataFrame:
    """Collapse duplicate-risk-factor rows by summing WS, keeping one row per
    distinct risk factor identity. Result preserves all schema fields needed
    by the rho_fn closures and _factor_label.

    Mathematically equivalent to the un-aggregated computation because for any
    two rows with the same identity, rho_kl = 1, so:
        WS_a^2 + WS_b^2 + 2*1*WS_a*WS_b = (WS_a + WS_b)^2
    """
    if rows.empty:
        return rows
    rows = rows.copy()
    rows["_fid"] = rows.apply(_factor_identity, axis=1)
    grouped = rows.groupby("_fid", as_index=False, sort=False).agg(
        id=("id", "first"),
        security=("security", "first"),
        instrument_type=("instrument_type", "first"),
        column=("column", "first"),
        kind=("kind", "first"),
        risk_class=("risk_class", "first"),
        bucket=("bucket", "first"),
        ccy=("ccy", "first"),
        name=("name", "first"),
        tenor=("tenor", "first"),
        under_tenor=("under_tenor", "first"),
        up_dn=("up_dn", "first"),
        value=("value", "sum"),
        ws=("ws", "sum"),
        n_position_rows=("ws", "count"),
    )
    grouped["factor_label"] = grouped.apply(_factor_label, axis=1)
    return grouped.drop(columns=["_fid"])


def compute_intra_bucket_with_trace(ws_df: pd.DataFrame, cfg: dict,
                                    scenario: str = "medium"):
    """Like compute_intra_bucket but additionally returns:
        position_factors_df  every original tidy row, grouped by bucket — for
                             tracing each Kb input back to a Stage 1 position
        distinct_factors_df  one row per distinct MAR21 risk factor (positions
                             with the same factor identity summed) — this is
                             what the rho matrix is built over
        rho_blocks           list of (bucket_label, factor_labels, rho_matrix)
                             one entry per multi-factor bucket; labels match
                             distinct_factors_df.factor_label
    """
    scen = scenario_fn(scenario)
    mask = ws_df["kind"].isin(["delta", "vega"])
    df = ws_df.loc[mask].copy()
    if df.empty:
        empty = pd.DataFrame(columns=["kind", "risk_class", "bucket_id",
                                      "n_distinct_factors", "n_position_rows",
                                      "Sb", "Kb", "mode", "scenario"])
        return empty, pd.DataFrame(), pd.DataFrame(), []

    def _bucket_id(row):
        if row["risk_class"] == "GIRR":
            return row["ccy"]
        if row["risk_class"] == "FX":
            return row["name"]
        return int(row["bucket"])
    df["bucket_id"] = df.apply(_bucket_id, axis=1)
    df["factor_label"] = df.apply(_factor_label, axis=1)

    summary_rows = []
    position_factor_rows = []
    distinct_factor_rows = []
    rho_blocks = []
    for (kind, rc, bid), sub in df.groupby(["kind", "risk_class", "bucket_id"]):
        sub = sub.reset_index(drop=True)
        Kb, Sb, n_distinct, rho_mat, mode, agg = \
            _aggregate_one_bucket_with_trace(sub, cfg, scen)
        summary_rows.append(dict(
            kind=kind, risk_class=rc, bucket_id=bid,
            n_distinct_factors=n_distinct, n_position_rows=len(sub),
            Sb=Sb, Kb=Kb, mode=mode, scenario=scenario))

        # Per-position trace: one row per original tidy row
        for _, row in sub.iterrows():
            position_factor_rows.append(dict(
                scenario=scenario, kind=kind, risk_class=rc, bucket_id=bid,
                factor_label=row["factor_label"],
                position_id=row["id"],
                column=row["column"], security=row["security"],
                tenor=row["tenor"], under_tenor=row["under_tenor"],
                value=row["value"],
                rw_base=row["rw_base"], divisor=row["divisor"],
                divisor_reference=row["divisor_reference"],
                rw=row["rw"], ws=row["ws"]))

        # Distinct-factor trace: one row per MAR21 risk factor (matches matrix)
        for _, row in agg.iterrows():
            distinct_factor_rows.append(dict(
                scenario=scenario, kind=kind, risk_class=rc, bucket_id=bid,
                factor_label=row["factor_label"],
                column=row["column"], security=row["security"],
                tenor=row["tenor"], under_tenor=row["under_tenor"],
                summed_value=row["value"],
                summed_ws=row["ws"],
                n_position_rows=row["n_position_rows"]))

        if mode == "pairwise" and n_distinct >= 2:
            labels = agg["factor_label"].tolist()
            rho_blocks.append((f"{rc} | {bid}", labels, rho_mat))

    summary_df = pd.DataFrame(summary_rows).sort_values(
        ["kind", "risk_class", "bucket_id"]).reset_index(drop=True)
    pos_df = pd.DataFrame(position_factor_rows)
    dist_df = pd.DataFrame(distinct_factor_rows)
    return summary_df, pos_df, dist_df, rho_blocks


def _write_rho_blocks_to_sheet(rho_blocks: list, sheet) -> None:
    """Write each (label, factor_labels, matrix) block stacked vertically into
    an openpyxl worksheet, with a heading row above each matrix."""
    row = 1
    for bucket_label, labels, mat in rho_blocks:
        sheet.cell(row=row, column=1,
                   value=f"Bucket: {bucket_label}   (n_factors={len(labels)})")
        row += 1
        # Header row: one blank cell then the column labels
        for j, lab in enumerate(labels):
            sheet.cell(row=row, column=j + 2, value=lab)
        row += 1
        for i, lab in enumerate(labels):
            sheet.cell(row=row, column=1, value=lab)
            for j in range(len(labels)):
                sheet.cell(row=row, column=j + 2, value=float(mat[i, j]))
            row += 1
        row += 2   # blank rows between blocks


if __name__ == "__main__":
    from openpyxl import Workbook

    print("Loading Phase 0 + Phase 1...")
    tidy = load_sensitivity_grid()
    cfg = load_config()
    ws_df = compute_weighted_sensitivities(tidy, cfg)

    for scen in ("medium", "high", "low"):
        print(f"\n=== PHASE 2 INTRA-BUCKET | scenario = {scen} ===")
        out = compute_intra_bucket(ws_df, cfg, scenario=scen)
        print(out.to_string(index=False,
                            formatters={"Sb": "{:,.2f}".format,
                                        "Kb": "{:,.2f}".format}))

    out_path = os.path.join(os.path.dirname(__file__), "output",
                            "phase2_intra_bucket.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # Build a self-describing README block for the workbook
    readme = pd.DataFrame([
        ("Sheet", "Purpose"),
        ("README", "This sheet — column / sheet glossary"),
        ("factors_distinct",
         "One row per MAR21 distinct risk factor in each bucket. summed_ws is "
         "the input to Kb. Row/column labels in the rho_matrices sheets match "
         "this sheet's factor_label exactly. SCENARIO-INDEPENDENT (WS = RW * value, "
         "neither depends on rho), so one sheet covers all three scenarios."),
        ("factors_per_position",
         "Every original tidy row, with full RW decomposition (rw_base, divisor, "
         "divisor_reference, rw) and the position id. Sum of ws within a "
         "factor_label equals the corresponding summed_ws in the distinct "
         "sheet — this is how you trace any Kb input back to a Stage 1 position. "
         "Also scenario-independent."),
        ("{scen}_summary",
         "One row per (kind, risk_class, bucket_id) with Sb, Kb, n_distinct_factors, "
         "n_position_rows, and mode (pairwise vs simple_sum). Kb depends on rho "
         "so this sheet IS scenario-specific."),
        ("{scen}_rho_matrices",
         "Stacked pairwise correlation matrices over DISTINCT risk factors, one "
         "block per multi-factor bucket. Single-factor buckets and 'simple_sum' "
         "MAR21 'Other' buckets are omitted (their Kb does not depend on rho). "
         "Matrices are scenario-scaled per MAR21.6, hence three sheets."),
        ("Kb formula",
         "Kb = sqrt( max(0, sum_k WS_k^2 + sum_{k != l} rho_kl * WS_k * WS_l) )  "
         "[MAR21.4(2)]. WS_k here is the summed_ws over distinct factors. Same-"
         "identity rows from multiple positions sum first (rho=1) -> identical "
         "Kb to the un-aggregated computation, but a much smaller, readable "
         "rho matrix for audit."),
    ], columns=["A", "B"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        readme.to_excel(xw, sheet_name="README", index=False, header=False)

        # Scenario-independent factor sheets — write once
        # (use medium scenario just to invoke the function; the dist/pos dfs are
        # identical across scenarios so the choice does not matter)
        _, pos_df, dist_df, _ = compute_intra_bucket_with_trace(
            ws_df, cfg, scenario="medium")
        # Drop the redundant 'scenario' column on the all-scenarios sheets
        dist_df.drop(columns=["scenario"], errors="ignore").to_excel(
            xw, sheet_name="factors_distinct", index=False)
        pos_df.drop(columns=["scenario"], errors="ignore").to_excel(
            xw, sheet_name="factors_per_position", index=False)

        for scen in ("medium", "high", "low"):
            summary, _, _, rho_blocks = compute_intra_bucket_with_trace(
                ws_df, cfg, scenario=scen)
            summary.to_excel(xw, sheet_name=f"{scen}_summary", index=False)
            ws_sheet = xw.book.create_sheet(f"{scen}_rho_matrices")
            _write_rho_blocks_to_sheet(rho_blocks, ws_sheet)
    print(f"\nWrote: {out_path}")
