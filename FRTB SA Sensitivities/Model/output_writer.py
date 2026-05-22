"""
FRTB SA — Module 4: Excel Output Writer
Formats the sensitivity DataFrame into a professional Excel file.
"""
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from mv_decompose import SCHEMA_COLUMNS as MV_SCHEMA_COLUMNS


def write_sensitivity_output(df: pd.DataFrame, mv_rows: list[dict], output_path: str):
    """Write sensitivity results to a formatted Excel file.

    Produces three sheets:
      * Sensitivities             (contract; byte-locked by qa_golden)
      * Portfolio_MV_Decomposed   (contract; model-priced MV blotter)
      * Summary                   (auxiliary; human review only)
    """
    
    FONT = 'Arial'
    H_FONT = Font(bold=True, size=9, color='FFFFFF', name=FONT)
    H_FILL = PatternFill('solid', fgColor='1F4E79')
    D_FONT = Font(size=8, name=FONT)
    FLAG_TRUE = PatternFill('solid', fgColor='E2EFDA')
    FLAG_FALSE = PatternFill('solid', fgColor='F2F2F2')
    SENS_FILL = PatternFill('solid', fgColor='FFF2CC')
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sensitivities"
    
    # Identify column groups
    meta_cols = ['Security', 'Instrument_Type', 'Sensitivity_Definition']
    flag_cols = [c for c in df.columns if c in [
        'GIRR_Delta', 'GIRR_Vega', 'GIRR_Curvature',
        'CSR_NonSec_Delta', 'CSR_Sec_Delta',
        'EQ_Delta', 'EQ_Vega', 'EQ_Curvature',
        'COMM_Delta', 'FX_Delta',
        'GIRR_Inflation', 'GIRR_XCcy_Basis',
    ]]
    sens_cols = [c for c in df.columns if c not in meta_cols + flag_cols
                 and c.startswith(('GIRR_', 'CSR_', 'EQ_', 'COMM_', 'FX_', 'VEGA_', 'CURV_'))]
    
    # Sort sensitivity columns by risk class, then by natural sort so that
    # every numeric segment (bucket IDs and tenors alike) sorts as a float,
    # not lexicographically.  'CSR_SEC_NONCTP_2_' < 'CSR_SEC_NONCTP_12_',
    # 'GIRR_GBP_2.0Y' < 'GIRR_GBP_10.0Y', etc.
    def sort_key(col):
        order = {'GIRR': 0, 'CSR_NONSEC': 1, 'CSR_SEC': 2, 'EQ': 3,
                 'VEGA': 4, 'CURV': 5, 'COMM': 6, 'FX': 7}
        # Split on every run of digits (with optional decimal), producing
        # alternating text / number segments.  Each segment becomes a
        # (text, float) 2-tuple so the overall key stays type-homogeneous
        # and Python's tuple comparison works without mixing str and float.
        parts = re.split(r'(\d+(?:\.\d+)?)', col)
        nat = tuple(
            ('', float(p)) if re.fullmatch(r'\d+(?:\.\d+)?', p) else (p, 0.0)
            for p in parts
        )
        for prefix, idx in order.items():
            if col.startswith(prefix):
                return (idx,) + nat
        return (99,) + nat
    
    sens_cols = sorted(sens_cols, key=sort_key)
    all_cols = meta_cols + flag_cols + sens_cols
    
    # Write headers
    # Row 1: Risk class group headers
    r = 1
    col_idx = 2  # Start after ID (col 1)
    
    # ID header
    cell = ws.cell(row=r, column=1, value="ID")
    cell.font = H_FONT; cell.fill = H_FILL; cell.border = thin
    
    for col_name in all_cols:
        col_idx += 0  # Will increment below
    
    # Row 2: Column headers
    r = 2
    ws.cell(row=r, column=1, value="ID").font = H_FONT
    ws.cell(row=r, column=1).fill = H_FILL
    ws.cell(row=r, column=1).border = thin
    
    for ci, col_name in enumerate(all_cols, 2):
        cell = ws.cell(row=r, column=ci, value=col_name)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', wrap_text=True, 
                                    text_rotation=90 if col_name in sens_cols else 0)
    
    # Write data
    for ri, (idx, row) in enumerate(df.iterrows(), 3):
        # ID
        cell = ws.cell(row=ri, column=1, value=idx)
        cell.font = Font(bold=True, size=9, name=FONT)
        cell.border = thin
        
        for ci, col_name in enumerate(all_cols, 2):
            val = row.get(col_name, '')
            cell = ws.cell(row=ri, column=ci)
            cell.border = thin
            cell.font = D_FONT
            
            if col_name in meta_cols:
                cell.value = str(val) if pd.notna(val) else ''
                cell.alignment = Alignment(wrap_text=True)
            elif col_name in flag_cols:
                cell.value = bool(val) if pd.notna(val) else False
                cell.fill = FLAG_TRUE if val else FLAG_FALSE
                cell.alignment = Alignment(horizontal='center')
            elif col_name in sens_cols:
                if pd.notna(val) and val != 0:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                    cell.fill = SENS_FILL
                else:
                    cell.value = 0
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    ws.column_dimensions['A'].width = 5  # ID
    for ci, col_name in enumerate(all_cols, 2):
        letter = get_column_letter(ci)
        if col_name in meta_cols:
            ws.column_dimensions[letter].width = 25 if 'Definition' in col_name else 18
        elif col_name in flag_cols:
            ws.column_dimensions[letter].width = 5
        else:
            ws.column_dimensions[letter].width = 5
    
    # Row heights
    ws.row_dimensions[2].height = 120  # Header row with rotated text
    for ri in range(3, 3 + len(df)):
        ws.row_dimensions[ri].height = 15
    
    # Freeze panes: freeze ID + meta columns, and header rows
    ws.freeze_panes = ws.cell(row=3, column=2 + len(meta_cols))

    # Sheet 2 — Portfolio_MV_Decomposed (contract sheet, inserted between
    # Sensitivities and Summary so the workbook order matches the contract).
    _write_portfolio_mv_decomposed(wb, mv_rows)

    # Add summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="FRTB SA Sensitivity Summary").font = Font(
        bold=True, size=14, name=FONT, color='1F4E79')
    ws2.cell(row=3, column=1, value="Total instruments").font = Font(bold=True, size=10, name=FONT)
    ws2.cell(row=3, column=2, value=len(df))
    
    # Count by type
    r = 5
    ws2.cell(row=r, column=1, value="By instrument type:").font = Font(bold=True, size=10, name=FONT)
    for itype, count in df['Instrument_Type'].value_counts().items():
        r += 1
        ws2.cell(row=r, column=1, value=itype)
        ws2.cell(row=r, column=2, value=count)
    
    r += 2
    ws2.cell(row=r, column=1, value="By risk class (True count):").font = Font(bold=True, size=10, name=FONT)
    for flag in flag_cols:
        r += 1
        ws2.cell(row=r, column=1, value=flag)
        ws2.cell(row=r, column=2, value=int(df[flag].sum()))
    
    r += 2
    ws2.cell(row=r, column=1, value="Active sensitivity columns:").font = Font(bold=True, size=10, name=FONT)
    active_sens = [c for c in sens_cols if df[c].abs().sum() > 0]
    ws2.cell(row=r, column=2, value=len(active_sens))
    for sc in active_sens:
        r += 1
        ws2.cell(row=r, column=1, value=sc)
        ws2.cell(row=r, column=2, value=df[sc].sum())
        ws2.cell(row=r, column=2).number_format = '#,##0.00'
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    wb.save(output_path)
    print(f"Output saved: {output_path}")
    print(f"  Sheet 1 'Sensitivities': {len(df)} instruments × {len(all_cols)} columns")
    print(f"  Sheet 2 'Portfolio_MV_Decomposed': {len(mv_rows)} legs × {len(MV_SCHEMA_COLUMNS)} columns")
    print(f"  {len(active_sens)} active sensitivity columns")


def _write_portfolio_mv_decomposed(wb: Workbook, mv_rows: list[dict]) -> None:
    """Write the Portfolio_MV_Decomposed sheet using the 14-column schema
    defined in `mv_decompose.SCHEMA_COLUMNS`.

    Strict ordering: rows are pre-sorted by the orchestrator via
    `mv_decompose.sort_rows`. Columns follow `SCHEMA_COLUMNS` exactly —
    no extra columns, no reordering.
    """
    FONT = 'Arial'
    H_FONT = Font(bold=True, size=9, color='FFFFFF', name=FONT)
    H_FILL = PatternFill('solid', fgColor='1F4E79')
    D_FONT = Font(size=9, name=FONT)
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))

    ws = wb.create_sheet('Portfolio_MV_Decomposed')

    for ci, col_name in enumerate(MV_SCHEMA_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for ri, leg in enumerate(mv_rows, 2):
        for ci, col_name in enumerate(MV_SCHEMA_COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=leg.get(col_name, ''))
            cell.font = D_FONT
            cell.border = thin
            if col_name in ('MV_USD', 'Quantity/Notional'):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_name in ('ID', 'parent_position_id'):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left', wrap_text=True)

    widths = {
        'ID': 10, 'parent_position_id': 14, 'decomposition_leg': 18,
        'Issuer': 18, 'Security': 36, 'Position Type': 12, 'Issue Type': 14,
        'Currency': 8, 'Long/Short': 9, 'Rating': 8, 'Seniority (DRC)': 14,
        'Quantity/Notional': 16, 'MV_USD': 16, 'pricing_model': 30,
    }
    for ci, col_name in enumerate(MV_SCHEMA_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col_name, 12)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = ws.cell(row=2, column=4)
